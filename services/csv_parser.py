import csv
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

SERIAL_TO_SENSOR = {
    "20536129-1": "ScaledVoltage",
    "20536129-2": "ScaledCurrent",
    "21433559-1": "ScaledSeries",
    "20466913-1": "TempIn1",
    "20466914-1": "TempIn2",
    "20466915-1": "TempOut1",
    "20590610-1": "TempOut2",
}
DEVICE_SERIAL = "21329018"

TZ = ZoneInfo("America/Toronto")
TWO_MINUTE_MS = 2 * 60 * 1000

PREFERRED_HEADER_HINTS = {
    "20536129-1": ["scaled voltage", ",vac", ",v,"],
    "20536129-2": ["scaled current"],
    "21433559-1": ["scaled series"],
    "20466913-1": ["temp in 1", "temp in"],
    "20466914-1": ["temp in 2", "temp in"],
    "20466915-1": ["temp out 1", "temp out"],
    "20590610-1": ["temp out 2", "temp out"],
}

EXCLUDED_HEADER_HINTS = {
    "20536129-1": ["voltage rms", "units"],
    "20536129-2": ["voltage rms"],
    "21433559-1": ["current ("],
}

def _parse_timestamp(date_value):
    if isinstance(date_value, datetime):
        dt = date_value if date_value.tzinfo else date_value.replace(tzinfo=TZ)
        return int(dt.timestamp() * 1000)

    date_text = str(date_value).strip()
    if not date_text:
        return None

    for fmt in ("%m/%d/%y %H:%M:%S %z", "%m/%d/%y %H:%M:%S"):
        try:
            dt = datetime.strptime(date_text, fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=TZ)
            return int(dt.timestamp() * 1000)
        except ValueError:
            continue
    return None


def _csv_rows(file_path):
    with open(file_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    header_idx = next((i for i, row in enumerate(rows) if "Date" in row), 0)
    headers = rows[header_idx]
    data_rows = rows[header_idx + 1 :]
    return headers, data_rows


def _xlsx_rows(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    data_rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    return headers, data_rows


def _get_headers_and_rows(file_path):
    if Path(file_path).suffix.lower() == ".xlsx":
        return _xlsx_rows(file_path)
    return _csv_rows(file_path)


def parse_licor_csv(csv_path):
    results = {}
    headers, rows = _get_headers_and_rows(csv_path)

    serial_to_indexes = {serial: [] for serial in SERIAL_TO_SENSOR.keys()}
    for idx, header in enumerate(headers):
        normalized = str(header or "").lower()
        for serial in serial_to_indexes:
            if serial in normalized:
                serial_to_indexes[serial].append(idx)

    # Rank candidate columns so the parser prefers expected scaled/Temp-in-out fields.
    ranked_indexes = {}
    for serial, indexes in serial_to_indexes.items():
        preferred = PREFERRED_HEADER_HINTS.get(serial, [])
        excluded = EXCLUDED_HEADER_HINTS.get(serial, [])

        def score(index):
            header_text = str(headers[index] or "").lower()
            if any(token in header_text for token in excluded):
                return 100
            if any(token in header_text for token in preferred):
                return 0
            return 10

        ranked_indexes[serial] = sorted(indexes, key=score)

    for serial in SERIAL_TO_SENSOR:
        results.setdefault(DEVICE_SERIAL, {})
        results[DEVICE_SERIAL].setdefault(serial, {"sensors": [{"data": [{"records": []}]}]})

    base_timestamp_ms = None
    sequence_index = 0

    for row in rows:
        if not row:
            continue

        date_idx = headers.index("Date") if "Date" in headers else None
        if date_idx is None or date_idx >= len(row):
            continue

        parsed_timestamp_ms = _parse_timestamp(row[date_idx])
        if parsed_timestamp_ms is None:
            continue
        if base_timestamp_ms is None:
            base_timestamp_ms = parsed_timestamp_ms

        # First pass: gather values present on this row.
        row_values = []
        for serial, indexes in ranked_indexes.items():
            value = None
            for idx in indexes:
                if idx < len(row):
                    candidate = row[idx]
                    if candidate not in (None, ""):
                        value = candidate
                        break

            if value is None:
                continue

            try:
                numeric_value = float(value)
            except (TypeError, ValueError):
                continue

            row_values.append((serial, numeric_value))

        # Skip rows that do not contain any target sensor values.
        # This prevents timeline drift when export files include empty/non-target rows.
        if not row_values:
            continue

        # Normalize imported file data to a strict 2-minute interval timeline.
        # Some exports contain irregular minute steps even though sampling is nominally 2 minutes.
        timestamp_ms = base_timestamp_ms + (sequence_index * TWO_MINUTE_MS)
        sequence_index += 1

        for serial, numeric_value in row_values:
            results[DEVICE_SERIAL][serial]["sensors"][0]["data"][0]["records"].append(
                [timestamp_ms, numeric_value]
            )

    return results